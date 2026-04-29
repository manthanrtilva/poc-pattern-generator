#!/usr/bin/env python3
"""Rainbow colors scrolling across 8 cells — output to Excel with background colors."""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

RAINBOW = [
    (0x94, 0x00, 0xD3),  # Violet
    (0x4B, 0x00, 0x82),  # Indigo
    (0x00, 0x00, 0xFF),  # Blue
    (0x00, 0xFF, 0x00),  # Green
    (0xFF, 0xFF, 0x00),  # Yellow
    (0xFF, 0x7F, 0x00),  # Orange
    (0xFF, 0x00, 0x00),  # Red
]

CELLS = 8

# Brightness ramp per cell: 0%→100% by 10%, then 100%→0% by 10%
BRIGHTNESS_CYCLE = list(range(0, 101, 10)) + list(range(90, 0, -10))
# = [0,10,20,...,100,90,80,...,10]  → 20 steps per cell

STEPS_PER_CELL = len(BRIGHTNESS_CYCLE)
CYCLES = 8
FRAMES = STEPS_PER_CELL * CYCLES  # 20 * 8 = 160


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Rainbow"

    # Set column widths for visual effect
    for col in range(1, CELLS + 1):
        ws.column_dimensions[chr(64 + col)].width = 12

    for row_idx in range(FRAMES):
        cycle = row_idx // STEPS_PER_CELL
        step = row_idx % STEPS_PER_CELL
        pct = BRIGHTNESS_CYCLE[step] / 100.0

        for col_idx in range(CELLS):
            base_r, base_g, base_b = RAINBOW[(col_idx + cycle) % len(RAINBOW)]
            r = int(base_r * pct)
            g = int(base_g * pct)
            b = int(base_b * pct)

            hex_color = f"{r:02X}{g:02X}{b:02X}"
            # Use white text on dark backgrounds, black text on light
            luminance = 0.299 * r + 0.587 * g + 0.114 * b
            font_color = "000000" if luminance > 128 else "FFFFFF"
            cell = ws.cell(row=row_idx + 1, column=col_idx + 1, value=f"#{hex_color}")
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            cell.font = Font(color=font_color)

    out = "rainbow.xlsx"
    wb.save(out)
    print(f"Wrote {out} ({FRAMES} rows x {CELLS} cols)")


if __name__ == "__main__":
    main()
